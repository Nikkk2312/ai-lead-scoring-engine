"""
Features 14-15, 32, 34: Excel writer with tier color-coding, summary stats, CSV export.
"""
import csv
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import OUTPUT_FILE, OUTPUT_CSV

TIER_FILLS = {
    "Hot": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
    "Warm": PatternFill(start_color="FFAA00", end_color="FFAA00", fill_type="solid"),
    "Cold": PatternFill(start_color="44BB44", end_color="44BB44", fill_type="solid"),
}
TIER_FONTS = {
    "Hot": Font(color="FFFFFF", bold=True),
    "Warm": Font(color="000000", bold=True),
    "Cold": Font(color="FFFFFF", bold=True),
}
HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

COLUMNS = [
    ("Company", "company_name", 22),
    ("Domain", "domain", 20),
    ("Score", "total_score", 10),
    ("Tier", "tier", 10),
    ("Confidence", "confidence", 12),
    ("Rule (/60)", "rule_score", 12),
    ("LLM (/40)", "soft_score", 12),
    ("Industry", "industry_classified", 24),
    ("HQ Country", "hq_country", 16),
    ("Founded", "founding_year", 12),
    ("Employees", "employee_estimate", 16),
    ("Tech Stack", "tech_stack", 35),
    ("Hiring Signal", "careers_jobs_count", 14),
    ("Competitors", "competitor_tech", 25),
    ("Key Signal", "key_signal", 30),
    ("Reasoning", "reasoning", 55),
    ("Outreach Line", "outreach_line", 45),
    ("Next Action", "next_action", 35),
    ("LinkedIn", "social_linkedin", 30),
    ("Email Pattern", "email_pattern", 22),
]


def write_results(scored_leads: list[dict], output_path: str | None = None) -> str:
    """Feature 14-15: Write scored leads to Excel with tier color-coding."""
    path = Path(output_path or OUTPUT_FILE)
    wb = Workbook()
    ws = wb.active
    ws.title = "Scored Leads"

    # Header row
    for col_idx, (header, _, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # Sort by tier then score
    tier_order = {"Hot": 0, "Warm": 1, "Cold": 2}
    sorted_leads = sorted(
        scored_leads,
        key=lambda x: (tier_order.get(x.get("tier", "Cold"), 3), -x.get("total_score", 0)),
    )

    # Data rows
    for row_idx, lead in enumerate(sorted_leads, 2):
        tier = lead.get("tier", "Cold")
        for col_idx, (_, key, _) in enumerate(COLUMNS, 1):
            value = lead.get(key)
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            elif isinstance(value, dict):
                value = json.dumps(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center",
                                       wrap_text=(key in ("reasoning", "key_signal", "outreach_line", "synthesis")))

        # Tier color-coding (Feature 15)
        for col in (3, 4):  # Score and Tier columns
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = TIER_FILLS.get(tier, PatternFill())
            cell.font = TIER_FONTS.get(tier, Font())
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Feature 32: Summary stats
    summary_row = len(sorted_leads) + 3
    ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True, size=12)

    hot_count = sum(1 for l in sorted_leads if l.get("tier") == "Hot")
    warm_count = sum(1 for l in sorted_leads if l.get("tier") == "Warm")
    cold_count = sum(1 for l in sorted_leads if l.get("tier") == "Cold")
    avg_score = sum(l.get("total_score", 0) for l in sorted_leads) / max(len(sorted_leads), 1)
    avg_conf = sum(l.get("confidence", 0) for l in sorted_leads) / max(len(sorted_leads), 1)

    stats = [
        ("Hot", hot_count, "Hot"),
        ("Warm", warm_count, "Warm"),
        ("Cold", cold_count, "Cold"),
        ("Total", len(sorted_leads), None),
        ("Avg Score", round(avg_score, 1), None),
        ("Avg Confidence", round(avg_conf, 2), None),
    ]

    for i, (label, value, tier) in enumerate(stats):
        row = summary_row + 1 + i
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        if tier:
            ws.cell(row=row, column=1).fill = TIER_FILLS.get(tier, PatternFill())
            ws.cell(row=row, column=1).font = TIER_FONTS.get(tier, Font())

    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"\n[OUTPUT] Excel written to: {path.resolve()}")
    return str(path.resolve())


def export_csv(scored_leads: list[dict], output_path: str | None = None) -> str:
    """Feature 34: CSV export of results."""
    path = Path(output_path or OUTPUT_CSV)
    csv_columns = [key for _, key, _ in COLUMNS]
    headers = [header for header, _, _ in COLUMNS]

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for lead in scored_leads:
            row = []
            for key in csv_columns:
                value = lead.get(key)
                if isinstance(value, (list, dict)):
                    value = json.dumps(value)
                row.append(value)
            writer.writerow(row)

    print(f"[OUTPUT] CSV written to: {path.resolve()}")
    return str(path.resolve())
